import sqlite3
import sys
import os
import datetime
from openpyxl import load_workbook


try:
    wb = load_workbook('config.xlsx')
except Exception:
    print('Error. config.xlsx has not been found')
    input('Press Enter')
    sys.exit()

try:
    sheet1 = wb['sheet1']
    sheet2 = wb['sheet2']
except Exception:
    print('Error. Sheets have not been found')
    input('Press Enter')
    sys.exit()
    
conn = sqlite3.connect(r'd:\Murs_smol_110\Database\Base_all.db3')
cursor = conn.cursor()

try:
    key2 = cursor.execute('SELECT * FROM OBJECTS;')
except Exception:
    conn.commit()
    conn.close()
    os.remove(r'd:\Murs_smol_110\Database\Base_all.db3')
    print('Error. Base_all.db3 has not been found')
    input('Press Enter')
    sys.exit()
  
i = 2

while (sheet1['A' + str(i)].value != None):

    if (sheet1['A' + str(i)].value == 4):
        shVal = sheet1['B' + str(i)].value
        if (shVal == None):
            break
        try:
            cursor.execute('DELETE FROM connections WHERE ID_CODE= (SELECT ID_CODE FROM objects WHERE CODE = "' + str(shVal) + '") OR ID_CONN =(SELECT ID_CODE FROM objects WHERE CODE = "' + str(shVal) +'");\n')
        except sqlite3.IntegrityError:
            print('Sheet1: Error in row #' + str(i))
            input('Press Enter')
            sys.exit()

    elif (sheet1['A' + str(i)].value == 1):
        shVal1 = sheet1['B' + str(i)].value
        shVal2 = sheet1['C' + str(i)].value
        if ((shVal1 == None) or (shVal2 == None)):
            break
        try:
            cursor.execute('INSERT OR REPLACE INTO connections VALUES((SELECT ID_CODE FROM objects WHERE CODE = "' + str(shVal1) + '"),(SELECT ID_CODE FROM objects WHERE CODE =  "' + str(shVal2) + '")),((SELECT ID_CODE FROM objects WHERE CODE =  "' + str(shVal2) + '"),(SELECT ID_CODE FROM objects WHERE CODE =  "' + str(shVal1) + '"));\n')
        except sqlite3.IntegrityError:
            print('Sheet1: Error in row #' + str(i))
            input('Press Enter')
            sys.exit()
            
    elif (sheet1['A' + str(i)].value == 2):
        shVal1 = sheet1['B' + str(i)].value
        shVal2 = sheet1['C' + str(i)].value
        if ((shVal1 == None) or (shVal2 == None)):
            break
        try:
           cursor.execute('DELETE FROM CONNECTIONS WHERE (ID_CODE = (SELECT ID_CODE FROM OBJECTS WHERE CODE = "' + str(shVal1) + '") AND ID_CONN = (SELECT ID_CODE FROM OBJECTS WHERE CODE = "' + str(shVal2) + '"))  OR  (ID_CODE = (SELECT ID_CODE FROM OBJECTS WHERE CODE = "' + str(shVal2) + '") AND   ID_CONN = (SELECT ID_CODE FROM OBJECTS WHERE CODE = "' + str(shVal1) + '"));\n')
        except sqlite3.OperationalError:
            print('Sheet1: Error in row #' + str(i))
            input('Press Enter')
            sys.exit()

    elif (sheet1['A' + str(i)].value == 3):
        shVal = sheet1['B' + str(i)].value
        if (shVal == None):
            break
        try:
            cursor.execute('UPDATE OBJECTS SET ID_TYPE=20 WHERE CODE =  "' + str(shVal) + '";\n')
        except sqlite3.OperationalError:
            print('Sheet1: Error in row #' + str(i))
            input('Press Enter')
            sys.exit()
        
    else:
        print('Sheet1: Error in cell [A' + str(i) + ']' )
        input('Press Enter')
        sys.exit()

    message1 = str (i) + ' rows from ' + str(sheet1.max_row)
    print(message1)
    i += 1

j = 2

while ((sheet2['A' + str(j)].value != None) and (sheet2['B' + str(j)].value != None)):
    shVal1 = sheet2['A' + str(j)].value
    shVal2 = sheet2['B' + str(j)].value
    try:
        cursor.execute('UPDATE OBJECTS SET FIDER10 =' + str(shVal2) + ' WHERE CODE = "' + str(shVal1) + '";\n')
    except sqlite3.OperationalError:
        print('Sheet2: Error in row #' + str(j))
        input('Press Enter')
        sys.exit()

    message2 = str (j) + ' rows from ' + str(sheet2.max_row)
    print(message2)
    j = j + 1
    
print('Sheet1: ' + message1 + ' have been finished')
print('Sheet2: ' + message2 + ' have been finished')

conn.commit()
conn.close()

input('Press Enter')
